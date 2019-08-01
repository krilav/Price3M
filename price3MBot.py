import datetime
import os
# import sqlite3

from openpyxl import load_workbook, Workbook
from pycbrf import ExchangeRates
import telebot
from telebot import apihelper
from datetime import timedelta

# from telebot import types

port = '7777'
usernameProxy = 'tg-id415061327'
passwordProxy = 'vjzTxv7v'
addressProxy = '@socksy.seriyps.ru'

apihelper.proxy = {'https': f'socks5://{usernameProxy}:{passwordProxy}{addressProxy}:{port}'}  # обход запрета в РФ

# используемые команды
commands = {
    'start ': 'Приветственная информация',
    'help  ': 'Показать информацию по вариантам команд',
    'config': 'Конфигурацию бота под пользователя'
}

TOKEN = '638610225:AAEoPelXhzUC11J11x8L9bBHbjoGPKj9zXk'  # Price3Mbot
# TOKEN = "842039603:AAFy4Cd_mWZSyjFEQGcUgI0uYP87ZrQy1pQ"  # pogodaDBbot
bot = telebot.TeleBot(TOKEN)

# параметры курса ЦБ
date_of_curse = str(datetime.date.today() - timedelta(1))
rates = ExchangeRates(date_of_curse, locale_en=True)
curse_eur_glob = float(rates['EUR'].value)


# Поиск по прайсу
def indexing(search):
    """Поиск нужных позиций."""

    file = 'db_1.xlsx'
    wb1 = load_workbook(filename=file)
    ws1 = wb1['price']

    # Поиск по ключам
    search = search.lower()
    search = search.split(' ')

    index = 0
    db_return = []

    for row_i in range(ws1.max_row - 1):
        tru_or_not_tru = []
        for search_i in search:
            tru_or_not_tru.append(search_i in ws1['w' + str(row_i + 2)].value)

        if False in tru_or_not_tru:
            pass
        else:
            index += 1
            name_of_search = ws1['f' + str(row_i + 2)].value
            cost_tmp = round(float(ws1['j' + str(row_i + 2)].value), 2)
            status_of_search = ws1['l' + str(row_i + 2)].value

            db_return_tmp = [index, name_of_search, cost_tmp, status_of_search]
            db_return.append(db_return_tmp)

    return db_return


# Команды '/help'.
@bot.message_handler(commands=['help'])
def command_help(message):
    help_text = "Возможные команды: \n"
    for key in commands:
        help_text += "/" + key + " :  "
        help_text += commands[key] + "\n"
    bot.send_message(message.chat.id, help_text)


# Команды '/start' и '/config'.
@bot.message_handler(commands=['start', 'config'])
def start_help(message):
    user_in_db = False
    if message.text == '/start':
        #        path_db_users = f'..{os.sep}Price3M{os.sep}db{os.sep}price3m.db'
        #        conn = sqlite3.connect(path_db_users)
        #        cursor = conn.cursor()
        #        cursor.execute("SELECT id FROM users")
        #        users_id = cursor.fetchall()

        #        for i in users_id:
        #            if message.from_user.id in i:
        #                user_in_db = True
        #                pass
        #            else:
        #                user_in_db = False

        #        if user_in_db:
        #            pass
        #        else:
        #            user = [(message.from_user.id, message.from_user.first_name, message.from_user.username,
        #                     message.from_user.last_name, message.from_user.language_code, '')
        #                    ]
        #            cursor.executemany("INSERT INTO users VALUES (?,?,?,?,?,?)", user)
        #            conn.commit()

        #        conn.close()

        bot.send_message(message.chat.id, '''Правила поиска по прайсу :
        1. Регистр не учитывается.
        2. Поиск идет по ключевым словам, которые нужно вводить через пробел.
        ''')
    elif message.text == '/config':
        bot.send_message(message.chat.id, f'''Типы цен :
        1. Прайсовая цена - цена без скидки.
        2. Оптовая цена - цена при обороте от 500€ в квартал (-15%)
        3. Крупнооптовая цена - цена при обороте от 1000€ в квартал (-29%)\n
         Расчетный курс € = {rates['EUR'].value}''')


# Основная реакция на ввод текста
@bot.message_handler(content_types=['text'])
def send_search_id(message):
    answer = indexing(message.text)
    # путь для exel
    file = f'{message.from_user.id}.xlsx'
    path_data_call = f'..{os.sep}Price3M{os.sep}data_call{os.sep}'
    # путь для db и запись
    #    path_db_users = f'..{os.sep}Price3M{os.sep}db{os.sep}price3m.db'
    #    conn = sqlite3.connect(path_db_users)
    #    cursor = conn.cursor()
    #    cursor.executemany("UPDATE users SET time_request = (?) where id = (?)",
    #                       [(str(datetime.datetime.today()), message.from_user.id)]
    #                       )
    #    conn.commit()

    if len(answer) > 3:
        bot.send_message(message.chat.id, 'Найдено более 3 наименований')
    elif len(answer) == 0:
        bot.send_message(message.chat.id, 'Не найдено ни одной позиции')
    # else:
    #    if file in os.listdir(path_data_call):
    #        wb = load_workbook(filename=f'{path_data_call}{file}')
    #        ws = wb.active
    #    else:
    #        wb = Workbook()
    #        ws = wb.active

    #    for answers in answer:
    #        ws.append(answers)
    #    wb.save(f'{path_data_call}{file}')

        for i in range(len(answer)):
            cost_send_bot = f'1. Прайсовая цена  -  {round(answer[i][2] * 1.2 * curse_eur_glob, 2)}'
            cost_send_bot += f'  руб. с НДС или {answer[i][2]}€  Без НДС'
            cost_send_bot += f'\n2. Оптовая цена  -  {round(answer[i][2] * 1.2 * curse_eur_glob * 0.85, 2)}'
            cost_send_bot += f'  руб. с НДС или {round(answer[i][2] * 0.85, 2)}€  Без НДС'
            cost_send_bot += f'\n3. Крупнооптовая цена  -  {round(answer[i][2] * 1.2 * curse_eur_glob * 0.71, 2)}'
            cost_send_bot += f'  руб. с НДС или {round(answer[i][2] * 0.71, 2)}€  Без НДC'

            bot.send_message(message.chat.id, f'№{answer[i][0]} {answer[i][1]}')
            bot.send_message(message.chat.id, cost_send_bot)
            bot.send_message(message.chat.id, f'Статус товара на складе 3М Россия - {answer[i][3]}')


bot.polling(none_stop=True)
