from openpyxl import load_workbook
from pycbrf import ExchangeRates


def new_data(curse_eur=0):
    """Дополнение даных расчетными параметрами."""

    for row_i in range(ws1.max_row - 1):

        try:
            # Добавление текущего курса
            _ = ws1.cell(column=15, row=row_i + 2, value="{0}".format(curse_eur))

            # Новое наименовение
            new_name_3m = ws1['f' + str(row_i + 2)].value.lower()
            _ = ws1.cell(column=23, row=row_i + 2, value="{0}".format(new_name_3m))

            # Разбиение статуса и срока
            word_left = {'дн.', '/', 'срок', 'поставки'}
            new_data_sk = list(set(ws1['l' + str(row_i + 2)].value.split(' ')) - word_left)
            new_data_sk.sort()

            if len(new_data_sk) == 1:
                _ = ws1.cell(column=16, row=row_i + 2, value="{0}".format(new_data_sk[0]))
            else:
                _ = ws1.cell(column=16, row=row_i + 2, value="{0}".format(new_data_sk[1]))
                _ = ws1.cell(column=17, row=row_i + 2, value="{0}".format(new_data_sk[0]))

            # Отбор ширины лент

            if ws1['f' + str(row_i + 2)].value.find('mm', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('mm', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('мм', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('мм', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('MM', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('MM', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('ММ', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('ММ', 1, 200)
            else:
                srz_start_point = 0

            print(srz_start_point)
            size_mm = ws1['f' + str(row_i + 2)].value[srz_start_point - 4:srz_start_point]
            _ = ws1.cell(column=19, row=row_i + 2, value="{0}".format(size_mm))
            size_mm.strip(' ')
            size_mm = int(size_mm.split(',')[0])

            if srz_start_point > 0:
                _ = ws1.cell(column=18, row=row_i + 2, value="{0}".format(size_mm))


        except:
            continue


# Разделение статуса и срока.

file = 'db_1.xlsx'
wb1 = load_workbook(filename=file)
ws1 = wb1['price']
rates = ExchangeRates('2019-07-03', locale_en=True)
curse_test_eur = rates['EUR'].value

new_data(curse_test_eur)

try:
    wb1.save(file)
except:
    print('!!!Открыт файл!!!')

