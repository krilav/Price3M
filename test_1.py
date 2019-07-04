from openpyxl import load_workbook
import time
from pycbrf import ExchangeRates

start_time = time.time()


# Дополнение даных расчетными параметрами


def new_data(curse_eur=0):
    """Дополнение даных расчетными параметрами."""

    for row_i in range(ws1.max_row - 1):

        try:
            # Добавление текущего курса
            _ = ws1.cell(column=15, row=row_i + 2, value="{0}".format(curse_eur))

            # Новое наименовение
            new_name_3m = ws1['f' + str(row_i + 2)].value.lower()
            _ = ws1.cell(column=23, row=row_i + 2, value="{0}".format(new_name_3m))

            # Разбиение статуса и срока
            word_left = {'дн.', '/', 'срок', 'поставки'}
            new_data_sk = list(set(ws1['l' + str(row_i + 2)].value.split(' ')) - word_left)
            new_data_sk.sort()

            if len(new_data_sk) == 1:
                _ = ws1.cell(column=16, row=row_i + 2, value="{0}".format(new_data_sk[0]))
            else:
                _ = ws1.cell(column=16, row=row_i + 2, value="{0}".format(new_data_sk[1]))
                _ = ws1.cell(column=17, row=row_i + 2, value="{0}".format(new_data_sk[0]))

            # Отбор ширины лент

            if ws1['f' + str(row_i + 2)].value.find('mm', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('mm', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('мм', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('мм', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('MM', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('MM', 1, 200)
            elif ws1['f' + str(row_i + 2)].value.find('ММ', 1, 200) > 0:
                srz_start_point = ws1['f' + str(row_i + 2)].value.find('ММ', 1, 200)
            else:
                srz_start_point = 0

            print(srz_start_point)
            size_mm = ws1['f' + str(row_i + 2)].value[srz_start_point - 4:srz_start_point]
            _ = ws1.cell(column=19, row=row_i + 2, value="{0}".format(size_mm))
            size_mm.strip(' ')
            size_mm = int(size_mm.split(',')[0])

            if srz_start_point > 0:
                _ = ws1.cell(column=18, row=row_i + 2, value="{0}".format(size_mm))


        except:
            continue


# Разделение статуса и срока.

def indexing(search='_'):
    """Поиск нужных позиций."""

    # Поиск по ключам
    search = search.split()

    number_of_index = 0
    list_of_search = []

    for row_i in range(ws1.max_row - 1):
        tru_or_not_tru = []
        for search_i in search:
            tru_or_not_tru.append(search_i in ws1['w' + str(row_i + 2)].value)

        if not False in tru_or_not_tru:
            _ = ws1.cell(column=14, row=row_i + 2, value="{0}".format(1))
            number_of_index += 1
            list_of_search.append(ws1['w' + str(row_i + 2)].value)
        else:
            _ = ws1.cell(column=14, row=row_i + 2, value="{0}".format(''))

    return list_of_search, number_of_index


file = 'db_1.xlsx'
wb1 = load_workbook(filename=file)
ws1 = wb1['price']
rates = ExchangeRates('2019-07-03', locale_en=True)
curse_test_eur = rates['EUR'].value

# ws2 = wb1.create_sheet(title='data2')

search_test = '550 600'

discount = 0.2

new_data(curse_test_eur)
# print(indexing(search_test)[0][0])

# print(ws1.max_row)
# print(ws1.max_column)

# size_mm_test = []
# for i in range(ws1.max_row-1):
#    srz_start_point = ws1['f' + str(i + 2)].value.find('мм х', 1, 100)
#    print(ws1['f' + str(i + 2)].value[srz_start_point-5:srz_start_point].strip(' '))
#    size_mm_test.append(ws1['f' + str(i + 2)].value[srz_start_point-5:srz_start_point])
# print(size_mm_test)

# srz_start_point = ws1['f9'].value.find('мм х', 1, 100)
# print(srz_start_point)
# print(ws1['f9'].value[srz_start_point - 5:srz_start_point])

# search = str(input('введите ключевое слово для поиска '))

try:
    wb1.save(file)
except:
    print('!!!Открыт файл!!!')

print(time.time() - start_time)
